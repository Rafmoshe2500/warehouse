from typing import List, Dict, Any, Optional, Tuple
import io
import logging
import re
import openpyxl
from app.db.mongodb import MongoDB
from app.services.bom_strategies import BomStrategyFactory

logger = logging.getLogger(__name__)


# ── Format Constants ──────────────────────────────────────────────────────────

FORMAT_NETAPP            = "netapp_pricing_template"
FORMAT_DELL              = "dell_quote"
FORMAT_HPE               = "hpe_quote"
FORMAT_CISCO             = "cisco_quote"
FORMAT_GENERIC_FIRST_COL = "generic_first_col"

SUPPORTED_FORMATS = {
    FORMAT_NETAPP,
    FORMAT_DELL,
    FORMAT_HPE,
    FORMAT_CISCO,
    FORMAT_GENERIC_FIRST_COL,
}

VALID_CATEGORIES = [
    "server-storage",
    "server",
    "switch",
    "io-card",
    "disk",
    "disk-shelf",
    "cable",
    "sfp-qsfp",
    "cpu",
    "memory",
    "fan",
    "psu",
    "license-capacity",
    "license-software",
    "support",
    "other",
]

# Categories that qualify as a top-level system (main item of a BOM group)
MAIN_CATEGORIES = {"server-storage", "disk-shelf", "switch"}


class BomService:
    def __init__(self):
        self.collection = MongoDB.get_collection("bom_part_catalog")

    # ── Group Reorganisation ──────────────────────────────────────────────────

    def _reorganize_groups(self, groups: List[Dict]) -> List[Dict]:
        """
        Ensure each group's main item belongs to MAIN_CATEGORIES.
        If it does not, scan children for a qualifying category and swap.
        Sets group["is_main_system"] = True/False on every group.
        """
        for group in groups:
            main_cat = group["main"].get("catalog", {}).get("category", "other")

            if main_cat not in MAIN_CATEGORIES:
                swap_idx = None
                for i, child in enumerate(group.get("children", [])):
                    child_cat = child.get("catalog", {}).get("category", "other")
                    if child_cat in MAIN_CATEGORIES:
                        swap_idx = i
                        break

                if swap_idx is not None:
                    old_main = group["main"]
                    group["main"] = group["children"][swap_idx]
                    group["children"][swap_idx] = old_main
                    logger.info(
                        "Reorganized BOM group: promoted child %s (cat: %s) to main, "
                        "demoted %s (cat: %s) to child",
                        group["main"].get("part_number"),
                        group["main"].get("catalog", {}).get("category"),
                        old_main.get("part_number"),
                        main_cat,
                    )

            final_cat = group["main"].get("catalog", {}).get("category", "other")
            group["is_main_system"] = final_cat in MAIN_CATEGORIES

        return groups

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _get_cell_value(self, row, col_index: Optional[int]):
        """Get cell value by 1-based column index."""
        if col_index is None:
            return None
        for cell in row:
            if cell.column == col_index:
                v = cell.value
                if isinstance(v, str):
                    return v.strip() or None
                return v
        return None

    def _parse_headers(self, ws, header_row: int, strategy) -> Dict[str, int]:
        """
        Return mapping of field_key → column index (1-based).
        Uses the strategy's column map with case-insensitive + whitespace-normalized matching.
        """
        col_map_source = strategy.get_column_map()

        def _normalize(s: str) -> str:
            return ' '.join(s.strip().split()).lower()

        normalized_source = {_normalize(k): v for k, v in col_map_source.items()}

        col_map = {}
        for cell in ws[header_row]:
            if not isinstance(cell.value, str):
                continue
            cell_norm = _normalize(cell.value)
            if not cell_norm:
                continue

            if cell_norm in normalized_source:
                col_map.setdefault(normalized_source[cell_norm], cell.column)
                continue

            for excel_norm, field_key in normalized_source.items():
                if excel_norm in cell_norm or cell_norm in excel_norm:
                    col_map.setdefault(field_key, cell.column)
                    break

        return col_map

    # ── Main Excel Parsing ────────────────────────────────────────────────────

    def parse_excel(self, file_bytes: bytes, fmt: str = FORMAT_NETAPP) -> Tuple[List[Dict], List[str], Dict[str, str]]:
        """
        Parse a BOM Excel file according to the given format strategy.

        Returns:
            groups: List of BOM group dicts (main + sub items)
            all_part_numbers: List of unique part numbers
            extracted_descriptions: Dict mapping part_number → product description (raw from Excel)
        """
        strategy = BomStrategyFactory.get_strategy(fmt)

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        groups: List[Dict] = []
        seen_part_numbers: set = set()
        extracted_descriptions: Dict[str, str] = {}

        # Iterate over every sheet (Dell has multiple sheets per file)
        for ws in wb.worksheets:
            header_row = strategy.find_header_row(ws)
            if not header_row:
                logger.debug("Skipping sheet '%s' — no header row found", ws.title)
                continue

            col_map = self._parse_headers(ws, header_row, strategy)
            if "part_number" not in col_map:
                logger.debug("Skipping sheet '%s' — no part_number column", ws.title)
                continue

            current_group: Optional[Dict] = None
            prev_part: Optional[str] = None

            for row in ws.iter_rows(min_row=header_row + 1):
                # Skip fully empty rows
                if not any(cell.value is not None for cell in row):
                    continue

                # Let the strategy filter out metadata rows (e.g. Cisco group banners)
                if hasattr(strategy, "is_data_row") and not strategy.is_data_row(row, col_map):
                    continue

                part_number = ""
                for cell in row:
                    if col_map.get("part_number") and cell.column == col_map["part_number"]:
                        if cell.value is not None:
                            part_number = str(cell.value).strip()
                        break

                is_header = strategy.is_group_header(row, col_map, prev_part)

                # Drop rows that are neither a group header nor carry a part number
                if not is_header and not part_number:
                    continue

                row_data = {
                    "part_number":    part_number,
                    "product":        self._get_cell_value(row, col_map.get("product")),
                    "ext_qty":        self._get_cell_value(row, col_map.get("ext_qty")),
                    "ext_list_price": self._get_cell_value(row, col_map.get("ext_list_price")),
                    "mod_group":      self._get_cell_value(row, col_map.get("mod_group")),
                    "net_discount":   self._get_cell_value(row, col_map.get("net_discount")),
                    "ext_net_price":  self._get_cell_value(row, col_map.get("ext_net_price")),
                }

                for field in ("ext_qty", "ext_list_price", "net_discount", "ext_net_price"):
                    try:
                        row_data[field] = float(row_data[field]) if row_data[field] is not None else 0.0
                    except (ValueError, TypeError):
                        row_data[field] = 0.0

                # "Qty N" in product description means each unit is a pack-of-N
                if row_data.get("product"):
                    m_qty = re.search(r'\bqty\s*(\d+)\b', str(row_data["product"]), re.I)
                    if m_qty:
                        row_data["ext_qty"] *= int(m_qty.group(1))

                if part_number:
                    seen_part_numbers.add(part_number)
                    if part_number not in extracted_descriptions and row_data["product"]:
                        extracted_descriptions[part_number] = str(row_data["product"]).strip()

                if is_header:
                    current_group = {
                        "main": row_data,
                        "children": [],
                        "total_net_price": 0.0,
                    }
                    groups.append(current_group)
                else:
                    if current_group is not None:
                        # Pull part number up to the parent header if missing (Dell pattern)
                        if not current_group["main"]["part_number"] and part_number:
                            current_group["main"]["part_number"] = part_number
                            if current_group["main"]["product"]:
                                extracted_descriptions[part_number] = str(current_group["main"]["product"]).strip()

                        current_group["children"].append(row_data)
                        if row_data["ext_net_price"] and row_data["ext_net_price"] > 0:
                            current_group["total_net_price"] += row_data["ext_net_price"]

                if part_number:
                    prev_part = part_number

        # Drop groups that ended up with no part number
        valid_groups = []
        for g in groups:
            if g["main"]["part_number"]:
                if not g["children"] and g["main"].get("ext_net_price"):
                    g["total_net_price"] = g["main"]["ext_net_price"]
                valid_groups.append(g)

        return valid_groups, list(seen_part_numbers), extracted_descriptions

    # ── Catalog Lookup ────────────────────────────────────────────────────────

    async def check_unknown_parts(self, part_numbers: List[str]) -> List[str]:
        """Return part numbers NOT found in bom_part_catalog."""
        if not part_numbers:
            return []
        cursor = self.collection.find(
            {"part_number": {"$in": part_numbers}},
            {"part_number": 1}
        )
        known = set()
        async for doc in cursor:
            known.add(doc["part_number"])
        return [p for p in part_numbers if p not in known]

    async def enrich_groups(self, groups: List[Dict]) -> List[Dict]:
        """Add catalog data to each item. For parts not in catalog, run AI classifier."""
        all_parts: Dict[str, str] = {}
        for group in groups:
            pn = group["main"].get("part_number")
            if pn:
                all_parts[pn] = group["main"].get("product", "")
            for child in group.get("children", []):
                cpn = child.get("part_number")
                if cpn:
                    all_parts[cpn] = child.get("product", "")

        if not all_parts:
            return groups

        catalog_map: Dict[str, Dict] = {}
        cursor = self.collection.find({"part_number": {"$in": list(all_parts.keys())}})
        async for doc in cursor:
            catalog_map[doc["part_number"]] = {
                "description_he": doc.get("description_he", ""),
                "category":       doc.get("category", "other"),
                "important":      doc.get("important", True),
            }

        unknown_pns = [pn for pn in all_parts if pn not in catalog_map]
        if unknown_pns:
            logger.info("Enrichment: %d unknown parts — triggering AI classifier", len(unknown_pns))
            try:
                from app.ai.classifier import classify_batch
                descriptions = [all_parts[pn] for pn in unknown_pns]
                ai_results = classify_batch(descriptions)
                for pn, desc, ai in zip(unknown_pns, descriptions, ai_results):
                    logger.info(
                        "[AI] %-35s | %3d%% | %-15s | %s",
                        pn[:35], int(ai["confidence"] * 100),
                        ai["category"], (desc or "")[:60],
                    )
                    catalog_map[pn] = {
                        "description_he": ai.get("description_he", ""),
                        "category":       ai.get("category", "other"),
                        "important":      True,
                        "_ai":            True,
                    }
            except Exception as exc:
                logger.warning("enrich_groups AI fallback: %s", exc)
                for pn in unknown_pns:
                    catalog_map[pn] = {"description_he": "", "category": "other", "important": True}

        for group in groups:
            pn = group["main"].get("part_number")
            group["main"]["catalog"] = catalog_map.get(pn, {"description_he": "", "category": "other", "important": True})
            for child in group.get("children", []):
                cpn = child.get("part_number")
                child["catalog"] = catalog_map.get(cpn, {"description_he": "", "category": "other", "important": True})

        return groups

    # ── Scan Entry Point ──────────────────────────────────────────────────────

    async def scan_bom(self, file_bytes: bytes, fmt: str = FORMAT_NETAPP) -> Dict:
        """
        Full BOM scan pipeline:
        1. Parse Excel using the given format strategy
        2. Check which parts are unknown in the catalog
        3. Enrich known parts with catalog data
        4. Run AI classifier on unknown parts (pre-fill suggestions)
        5. Return result
        """
        logger.info("Starting BOM scan — format: %s  size: %d bytes", fmt, len(file_bytes))
        groups, all_part_numbers, extracted_descriptions = self.parse_excel(file_bytes, fmt)
        unknown_parts = await self.check_unknown_parts(all_part_numbers)
        enriched_groups = await self.enrich_groups(groups)
        enriched_groups = self._reorganize_groups(enriched_groups)

        unknown_list = [
            {"part_number": pn, "excel_description": extracted_descriptions.get(pn, "")}
            for pn in unknown_parts if pn
        ]

        if unknown_list:
            try:
                from app.ai.classifier import classify_batch
                descriptions = [item["excel_description"] for item in unknown_list]
                ai_results = classify_batch(descriptions)
                for item, ai in zip(unknown_list, ai_results):
                    item["ai_label"]          = ai["label"]
                    item["ai_category"]       = ai["category"]
                    item["ai_description_he"] = ai["description_he"]
                    item["ai_confidence"]     = ai["confidence"]
                    item["ai_low_confidence"] = ai["low_confidence"]
                    item["ai_attributes"]     = ai.get("attributes", {})
            except Exception as exc:
                logger.warning("AI classification skipped or failed: %s", exc)

        return {
            "groups":        enriched_groups,
            "unknown_parts": unknown_list,
            "total_groups":  len(enriched_groups),
        }

    # ── Catalog CRUD ──────────────────────────────────────────────────────────

    async def save_part(
        self,
        part_number: str,
        description_he: str,
        category: str,
        important: bool,
        excel_description: str = "",
    ) -> Dict:
        """Upsert a part into bom_part_catalog."""
        from datetime import datetime, timezone

        if category not in VALID_CATEGORIES:
            raise ValueError(f"קטגוריה לא חוקית: {category}")

        doc = {
            "part_number":       part_number,
            "description_he":    description_he,
            "category":          category,
            "important":         important,
            "excel_description": excel_description,
            "updated_at":        datetime.now(timezone.utc).isoformat(),
        }

        await self.collection.update_one(
            {"part_number": part_number},
            {
                "$set": doc,
                "$setOnInsert": {"created_at": datetime.now(timezone.utc).isoformat()},
            },
            upsert=True,
        )
        return doc

    async def get_all_parts(self) -> List[Dict]:
        """Return all parts in the catalog."""
        parts = []
        cursor = self.collection.find({}, {"_id": 0}).sort("part_number", 1)
        async for doc in cursor:
            parts.append(doc)
        return parts