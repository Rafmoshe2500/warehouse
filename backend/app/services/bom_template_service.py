"""
BomTemplateService — business logic for managing BOM template configurations.
"""
import io
import re
import logging
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional

import openpyxl

from app.db.repositories.bom_template_repository import BomTemplateRepository
from app.services.bom_strategies.dynamic_strategy import DynamicBomStrategy

logger = logging.getLogger(__name__)


def _slugify(name: str) -> str:
    """Turn a vendor display name into a URL/DB-safe slug."""
    slug = name.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    return slug.strip("_")


class BomTemplateService:
    def __init__(self):
        self.repo = BomTemplateRepository()

    # ── CRUD ──────────────────────────────────────────────────────────────────

    async def create_template(self, data: dict, username: str) -> dict:
        format_id = _slugify(data["vendor_name"])
        if not format_id:
            raise ValueError("vendor_name must produce a valid slug")

        existing = await self.repo.get_by_format_id(format_id)
        if existing:
            raise ValueError(f"Template with format_id '{format_id}' already exists")

        now = datetime.now(timezone.utc)
        doc = {
            "vendor_name":      data["vendor_name"],
            "format_id":        format_id,
            "description":      data.get("description"),
            "column_map":       data["column_map"],
            "header_detection": data["header_detection"],
            "group_detection":  data["group_detection"],
            "data_row_filter":  data.get("data_row_filter"),
            "color":            data.get("color"),
            "logo":             data.get("logo"),
            "is_active":        True,
            "is_builtin":       False,
            "created_by":       username,
            "created_at":       now,
            "updated_at":       now,
        }
        created = await self.repo.create(doc)
        logger.info("BOM template created: format_id=%s by %s", format_id, username)
        return self.repo._format_doc(created)

    async def update_template(self, template_id: str, data: dict, username: str) -> Optional[dict]:
        existing = await self.repo.get_by_id(template_id)
        if not existing:
            return None

        update_fields: Dict[str, Any] = {}
        for key in (
            "vendor_name", "description", "column_map",
            "header_detection", "group_detection", "data_row_filter",
            "is_active", "color", "logo",
        ):
            if key in data and data[key] is not None:
                update_fields[key] = data[key]

        if not update_fields:
            return self.repo._format_doc(existing)

        update_fields["updated_at"] = datetime.now(timezone.utc)
        updated = await self.repo.update(template_id, update_fields)
        logger.info("BOM template updated: id=%s by %s", template_id, username)
        return self.repo._format_doc(updated)

    async def delete_template(self, template_id: str) -> Optional[dict]:
        existing = await self.repo.get_by_id(template_id)
        if not existing:
            return None
        if existing.get("is_builtin"):
            return await self.repo.deactivate(template_id)
        return await self.repo.deactivate(template_id)

    async def get_all_templates(self) -> List[dict]:
        docs = await self.repo.get_active_templates()
        return [self.repo._format_doc(d) for d in docs]

    async def get_template(self, template_id: str) -> Optional[dict]:
        doc = await self.repo.get_by_id(template_id)
        return self.repo._format_doc(doc) if doc else None

    # ── Excel preview (for wizard UI) ─────────────────────────────────────────

    @staticmethod
    def preview_excel(file_bytes: bytes, max_rows: int = 50) -> dict:
        """Return the first N rows of the first sheet as JSON for the wizard UI."""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.worksheets[0]

        rows = []
        for idx, row in enumerate(ws.iter_rows(max_row=max_rows), start=1):
            cells = []
            for cell in row:
                cell_data: Dict[str, Any] = {
                    "value": cell.value if not isinstance(cell.value, datetime) else cell.value.isoformat(),
                    "col":   cell.column,
                }
                # Include fill colour info so the wizard can show colour fills
                try:
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.type == "rgb" and fill.fgColor.rgb:
                        cell_data["fill"] = fill.fgColor.rgb.upper()
                except Exception:
                    pass
                cells.append(cell_data)
            rows.append({"row": idx, "cells": cells})

        return {
            "sheet_name": ws.title,
            "total_sheets": len(wb.worksheets),
            "rows": rows,
        }

    # ── Validate template config against a sample file ────────────────────────

    @staticmethod
    def validate_template(template_data: dict, file_bytes: bytes) -> dict:
        """Parse a sample file using the given template config and return a preview."""
        strategy = DynamicBomStrategy(template_data)

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.worksheets[0]

        header_row = strategy.find_header_row(ws)
        if header_row is None:
            return {
                "valid": False,
                "error": f"Header row not found — keyword '{template_data['header_detection']['keyword']}' not detected in first {template_data['header_detection'].get('max_scan_rows', 25)} rows",
                "groups_count": 0,
                "sample_group": None,
            }

        # Build col_map
        col_map_source = strategy.get_column_map()

        def _normalize(s: str) -> str:
            return " ".join(s.strip().split()).lower()

        normalized_source = {_normalize(k): v for k, v in col_map_source.items()}
        col_map: Dict[str, int] = {}
        for cell in ws[header_row]:
            if not isinstance(cell.value, str):
                continue
            cell_norm = _normalize(cell.value)
            if not cell_norm:
                continue
            if cell_norm in normalized_source:
                col_map.setdefault(normalized_source[cell_norm], cell.column)
                continue
            for excel_norm, field_key in normalized_source.items():
                if excel_norm in cell_norm or cell_norm in excel_norm:
                    col_map.setdefault(field_key, cell.column)
                    break

        if "part_number" not in col_map:
            return {
                "valid": False,
                "error": "Could not map 'part_number' column from the header row",
                "groups_count": 0,
                "sample_group": None,
            }

        # Parse groups
        groups = []
        current_group = None
        prev_part = None

        for row in ws.iter_rows(min_row=header_row + 1):
            if not any(cell.value is not None for cell in row):
                continue
            if hasattr(strategy, "is_data_row") and not strategy.is_data_row(row, col_map):
                continue

            pn = ""
            for cell in row:
                if col_map.get("part_number") and cell.column == col_map["part_number"]:
                    if cell.value is not None:
                        pn = str(cell.value).strip()
                    break

            is_header = strategy.is_group_header(row, col_map, prev_part)
            if not is_header and not pn:
                continue

            product_val = None
            product_col = col_map.get("product")
            if product_col:
                for cell in row:
                    if cell.column == product_col:
                        product_val = str(cell.value).strip() if cell.value else None
                        break

            if is_header:
                current_group = {"main": pn, "main_product": product_val, "children": []}
                groups.append(current_group)
            elif current_group is not None:
                if pn:
                    current_group["children"].append({"part_number": pn, "product": product_val})

            if pn:
                prev_part = pn

        sample = groups[0] if groups else None

        return {
            "valid": True,
            "error": None,
            "header_row": header_row,
            "mapped_columns": list(col_map.keys()),
            "groups_count": len(groups),
            "total_children": sum(len(g["children"]) for g in groups),
            "sample_group": sample,
        }
