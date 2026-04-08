"""
Tests for NetAppBomStrategy using real example BOM file (דוגמא.xlsx).
"""
import os
import pytest
import openpyxl
from app.services.bom_strategies.netapp_strategy import NetAppBomStrategy

EXAMPLE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "example_bom_files", "דוגמא.xlsx"
)


@pytest.fixture
def strategy():
    return NetAppBomStrategy()


@pytest.fixture
def workbook():
    assert os.path.exists(EXAMPLE_FILE), f"Example BOM file not found: {EXAMPLE_FILE}"
    return openpyxl.load_workbook(EXAMPLE_FILE, data_only=True)


class TestNetAppStrategyProperties:

    def test_format_id(self, strategy):
        assert strategy.format_id == "netapp_pricing_template"

    def test_column_map_keys(self, strategy):
        col_map = strategy.get_column_map()
        assert "Part Number" in col_map
        assert col_map["Part Number"] == "part_number"
        assert "Product" in col_map
        assert col_map["Product"] == "product"
        assert "Ext Qty" in col_map
        assert "Ext Net Price" in col_map
        assert "Net Discount" in col_map


class TestNetAppFindHeaderRow:

    def test_finds_header_in_real_file(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        assert header_row is not None
        assert isinstance(header_row, int)
        assert header_row >= 1

    def test_header_row_contains_part_number(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        found = False
        for cell in ws[header_row]:
            if isinstance(cell.value, str) and "part number" in cell.value.lower():
                found = True
                break
        assert found, f"Row {header_row} does not contain 'Part Number' header"

    def test_returns_none_for_empty_sheet(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert strategy.find_header_row(ws) is None


class TestNetAppIsGroupHeader:

    def test_yellow_row_is_group_header(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        assert header_row is not None

        # Build col_map from header row
        col_map = {}
        source = strategy.get_column_map()
        for cell in ws[header_row]:
            if isinstance(cell.value, str):
                for excel_key, field_key in source.items():
                    if cell.value.strip().lower() == excel_key.lower():
                        col_map[field_key] = cell.column

        assert "part_number" in col_map, "Could not map part_number column"

        # Find at least one group header row (yellow) in the file
        group_header_found = False
        for row in ws.iter_rows(min_row=header_row + 1):
            if strategy.is_group_header(row, col_map):
                group_header_found = True
                break

        assert group_header_found, "No group header (yellow row) found in NetApp example"

    def test_non_yellow_row_is_not_group_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="PART-123")
        col_map = {"part_number": 1}
        row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        assert strategy.is_group_header(row, col_map) is False


class TestNetAppFullParseCycle:
    """Integration-level check: strategy + BomService together parse the real file."""

    def test_parse_produces_groups(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        assert header_row is not None

        col_map = {}
        source = strategy.get_column_map()
        for cell in ws[header_row]:
            if isinstance(cell.value, str):
                for excel_key, field_key in source.items():
                    if cell.value.strip().lower() == excel_key.lower():
                        col_map.setdefault(field_key, cell.column)

        groups_count = 0
        children_count = 0
        current_group = None

        for row in ws.iter_rows(min_row=header_row + 1):
            if not any(cell.value is not None for cell in row):
                continue
            if strategy.is_group_header(row, col_map):
                groups_count += 1
                current_group = True
            elif current_group:
                pn = None
                for cell in row:
                    if col_map.get("part_number") and cell.column == col_map["part_number"]:
                        pn = cell.value
                if pn:
                    children_count += 1

        assert groups_count > 0, "Expected at least one group in NetApp example"
        assert children_count > 0, "Expected at least one child item in NetApp example"
