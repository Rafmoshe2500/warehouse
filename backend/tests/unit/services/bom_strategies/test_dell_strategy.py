"""
Tests for DellBomStrategy using real example BOM file (דוגמא2.xlsx).
"""
import os
import pytest
import openpyxl
from app.services.bom_strategies.dell_strategy import DellBomStrategy

EXAMPLE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "example_bom_files", "דוגמא2.xlsx"
)


@pytest.fixture
def strategy():
    return DellBomStrategy()


@pytest.fixture
def workbook():
    assert os.path.exists(EXAMPLE_FILE), f"Example BOM file not found: {EXAMPLE_FILE}"
    return openpyxl.load_workbook(EXAMPLE_FILE, data_only=True)


class TestDellStrategyProperties:

    def test_format_id(self, strategy):
        assert strategy.format_id == "dell_quote"

    def test_column_map_handles_sku_variants(self, strategy):
        col_map = strategy.get_column_map()
        assert col_map["Sku"] == "part_number"
        assert col_map["SKU"] == "part_number"

    def test_column_map_handles_qty_variants(self, strategy):
        col_map = strategy.get_column_map()
        assert col_map["Qty"] == "ext_qty"
        assert col_map["Quantity"] == "ext_qty"

    def test_column_map_handles_price_variants(self, strategy):
        col_map = strategy.get_column_map()
        assert col_map["Total Selling Price"] == "ext_net_price"
        assert col_map["TotalSelling Price"] == "ext_net_price"
        assert col_map["Total List Price"] == "ext_list_price"
        assert col_map["TotalList Price"] == "ext_list_price"

    def test_column_map_has_category(self, strategy):
        col_map = strategy.get_column_map()
        assert col_map["Category"] == "mod_group"

    def test_column_map_has_unit_prices(self, strategy):
        col_map = strategy.get_column_map()
        assert "Unit List Price" in col_map or "UnitList Price" in col_map
        assert "Unit Selling Price" in col_map or "UnitSelling Price" in col_map


class TestDellFindHeaderRow:

    def test_finds_header_in_real_file(self, strategy, workbook):
        found = False
        for ws in workbook.worksheets:
            header_row = strategy.find_header_row(ws)
            if header_row is not None:
                found = True
                break
        assert found, "No header row found in any sheet of Dell example"

    def test_header_row_contains_sku(self, strategy, workbook):
        for ws in workbook.worksheets:
            header_row = strategy.find_header_row(ws)
            if header_row is not None:
                sku_found = False
                for cell in ws[header_row]:
                    if isinstance(cell.value, str) and "sku" in cell.value.lower():
                        sku_found = True
                        break
                assert sku_found, f"Row {header_row} does not contain 'Sku' header"
                break

    def test_returns_none_for_empty_sheet(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert strategy.find_header_row(ws) is None


class TestDellIsGroupHeader:

    def test_yellow_sku_cell_is_group_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="DELL-SYS")
        cell.fill = openpyxl.styles.PatternFill(
            start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
        )
        col_map = {"part_number": 1}
        row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        assert strategy.is_group_header(row, col_map) is True

    def test_yellow_non_sku_cell_is_group_header(self, strategy):
        """Dell: any yellow cell with value in the row triggers group header."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="DELL-X")
        cell2 = ws.cell(row=1, column=2, value="Server Description")
        cell2.fill = openpyxl.styles.PatternFill(
            start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
        )
        col_map = {"part_number": 1}
        row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        assert strategy.is_group_header(row, col_map) is True

    def test_non_yellow_row_is_not_group_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="DELL-CHILD")
        col_map = {"part_number": 1}
        row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        assert strategy.is_group_header(row, col_map) is False

    def test_real_file_has_group_headers(self, strategy, workbook):
        found = False
        for ws in workbook.worksheets:
            header_row = strategy.find_header_row(ws)
            if header_row is None:
                continue
            source = strategy.get_column_map()
            col_map = {}
            for cell in ws[header_row]:
                if isinstance(cell.value, str):
                    for excel_key, field_key in source.items():
                        if cell.value.strip().lower() == excel_key.lower():
                            col_map.setdefault(field_key, cell.column)
            for row in ws.iter_rows(min_row=header_row + 1):
                if strategy.is_group_header(row, col_map):
                    found = True
                    break
            if found:
                break
        assert found, "No group headers found in Dell example file"


class TestDellMultipleSheets:
    """Dell files typically have multiple sheets."""

    def test_workbook_has_sheets(self, workbook):
        assert len(workbook.worksheets) >= 1

    def test_at_least_one_sheet_has_data(self, strategy, workbook):
        sheets_with_headers = 0
        for ws in workbook.worksheets:
            if strategy.find_header_row(ws) is not None:
                sheets_with_headers += 1
        assert sheets_with_headers >= 1, "No parseable sheets found in Dell example"
