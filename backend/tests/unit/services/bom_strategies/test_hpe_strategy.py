"""
Tests for HpeBomStrategy using real example BOM file (דוגמא3.xlsx).
"""
import os
import pytest
import openpyxl
from app.services.bom_strategies.hpe_strategy import HpeBomStrategy

EXAMPLE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "example_bom_files", "דוגמא3.xlsx"
)


@pytest.fixture
def strategy():
    return HpeBomStrategy()


@pytest.fixture
def workbook():
    assert os.path.exists(EXAMPLE_FILE), f"Example BOM file not found: {EXAMPLE_FILE}"
    return openpyxl.load_workbook(EXAMPLE_FILE, data_only=True)


class TestHpeStrategyProperties:

    def test_format_id(self, strategy):
        assert strategy.format_id == "hpe_quote"

    def test_column_map_keys(self, strategy):
        col_map = strategy.get_column_map()
        assert col_map["UCID"] == "part_number"
        assert col_map["Description"] == "product"
        assert col_map["Qty"] == "ext_qty"
        assert col_map["Unit List Price"] == "unit_list_price"
        assert col_map["Total List Price"] == "ext_list_price"
        assert col_map["Unit Net Price"] == "unit_net_price"
        assert col_map["Total Net Price"] == "ext_net_price"


class TestHpeFindHeaderRow:

    def test_finds_header_in_real_file(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        assert header_row is not None
        assert isinstance(header_row, int)

    def test_header_row_contains_ucid(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        if header_row is not None:
            found = False
            for cell in ws[header_row]:
                if isinstance(cell.value, str) and "ucid" in cell.value.lower():
                    found = True
                    break
            assert found, f"Row {header_row} does not contain 'UCID' header"

    def test_returns_none_for_empty_sheet(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert strategy.find_header_row(ws) is None


class TestHpeIsGroupHeader:

    def test_every_row_is_group_header(self, strategy):
        """HPE strategy marks every row as a group header (flat structure)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="HPE-PART")
        col_map = {"part_number": 1}
        row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        assert strategy.is_group_header(row, col_map) is True

    def test_empty_row_is_still_group_header(self, strategy):
        """HPE strategy returns True unconditionally."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=None)
        col_map = {}
        row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        assert strategy.is_group_header(row, col_map) is True


class TestHpeFullParseCycle:

    def test_real_file_has_data_rows(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        assert header_row is not None

        col_map = {}
        source = strategy.get_column_map()
        for cell in ws[header_row]:
            if isinstance(cell.value, str):
                for excel_key, field_key in source.items():
                    if cell.value.strip().lower() == excel_key.lower():
                        col_map.setdefault(field_key, cell.column)

        data_rows = 0
        for row in ws.iter_rows(min_row=header_row + 1):
            if any(cell.value is not None for cell in row):
                data_rows += 1
        assert data_rows > 0, "No data rows found in HPE example"

    def test_all_rows_become_groups(self, strategy, workbook):
        """Because is_group_header always returns True, every row becomes a group."""
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        if header_row is None:
            pytest.skip("No header row found in HPE example")

        col_map = {}
        source = strategy.get_column_map()
        for cell in ws[header_row]:
            if isinstance(cell.value, str):
                for excel_key, field_key in source.items():
                    if cell.value.strip().lower() == excel_key.lower():
                        col_map.setdefault(field_key, cell.column)

        groups = 0
        for row in ws.iter_rows(min_row=header_row + 1):
            if any(cell.value is not None for cell in row):
                if strategy.is_group_header(row, col_map):
                    groups += 1
        assert groups > 0
