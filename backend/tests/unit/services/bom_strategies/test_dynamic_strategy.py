"""
Tests for DynamicBomStrategy – group detection modes, header detection, data row filtering.
"""
import pytest
from openpyxl import Workbook
from app.services.bom_strategies.dynamic_strategy import DynamicBomStrategy


def _make_config(**overrides):
    base = {
        "format_id": "test_vendor",
        "vendor_name": "TestVendor",
        "header_detection": {
            "keyword": "part number",
            "max_scan_rows": 20,
        },
        "column_map": {
            "Part Number": "part_number",
            "Description": "product",
            "Qty": "ext_qty",
        },
        "group_detection": {
            "mode": "all_rows",
            "config": {},
        },
    }
    base.update(overrides)
    return base


class TestDynamicBomStrategyInit:

    def test_creates_with_valid_config(self):
        config = _make_config()
        strategy = DynamicBomStrategy(config)
        assert strategy.format_id == "test_vendor"

    def test_format_id_property(self):
        config = _make_config(format_id="custom_fmt")
        strategy = DynamicBomStrategy(config)
        assert strategy.format_id == "custom_fmt"

    def test_get_column_map_returns_copy(self):
        config = _make_config()
        strategy = DynamicBomStrategy(config)
        cm = strategy.get_column_map()
        assert cm == {"Part Number": "part_number", "Description": "product", "Qty": "ext_qty"}


class TestDynamicBomStrategyHeader:

    def test_find_header_row_by_keyword(self):
        config = _make_config()
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["Intro row"])
        ws.append(["Part Number", "Description", "Qty"])
        ws.append(["ABC-123", "Widget", 5])
        header_row = strategy.find_header_row(ws)
        assert header_row == 2

    def test_find_header_row_case_insensitive(self):
        config = _make_config()
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["PART NUMBER", "Desc"])
        assert strategy.find_header_row(ws) == 1

    def test_find_header_row_not_found(self):
        config = _make_config()
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["Nothing here"])
        ws.append(["Also nothing"])
        header_row = strategy.find_header_row(ws)
        assert header_row is None


class TestDynamicBomStrategyGroupDetection:

    def test_all_rows_mode_returns_true(self):
        config = _make_config(group_detection={"mode": "all_rows", "config": {}})
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["A-001", "Item A", 1])
        # is_group_header takes (row, col_map, prev_part)
        col_map = {"part_number": 1, "product": 2, "ext_qty": 3}
        assert strategy.is_group_header(ws[1], col_map) is True

    def test_unknown_mode_returns_false(self):
        config = _make_config(group_detection={"mode": "nonexistent", "config": {}})
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["A-001", "Item", 1])
        col_map = {"part_number": 1}
        assert strategy.is_group_header(ws[1], col_map) is False


class TestDynamicBomStrategyDataRow:

    def test_no_filter_accepts_all(self):
        config = _make_config()
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["anything", "here"])
        col_map = {"part_number": 1}
        assert strategy.is_data_row(ws[1], col_map) is True

    def test_filter_pattern_match_accepts_row(self):
        config = _make_config()
        config["data_row_filter"] = {"column": "part_number", "pattern": r"^\d+\.\d+$"}
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["1.2", "Item"])
        col_map = {"part_number": 1}
        assert strategy.is_data_row(ws[1], col_map) is True

    def test_filter_pattern_no_match_rejects_row(self):
        config = _make_config()
        config["data_row_filter"] = {"column": "part_number", "pattern": r"^\d+\.\d+$"}
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["skip_me", "Item"])
        col_map = {"part_number": 1}
        assert strategy.is_data_row(ws[1], col_map) is False

    def test_filter_column_not_in_col_map_accepts_all(self):
        config = _make_config()
        config["data_row_filter"] = {"column": "nonexistent_col", "pattern": r"^\d+$"}
        strategy = DynamicBomStrategy(config)
        wb = Workbook()
        ws = wb.active
        ws.append(["anything"])
        col_map = {"part_number": 1}
        assert strategy.is_data_row(ws[1], col_map) is True


# ── Helper: MockCell ──────────────────────────────────────────────────────────


def make_cell(col: int, value=None, rgb: str = None):
    """Create a MagicMock cell that simulates openpyxl cell behaviour."""
    from unittest.mock import MagicMock
    cell = MagicMock()
    cell.column = col
    cell.value = value
    fill = MagicMock()
    fill.fgColor.type = "rgb" if rgb else "none"
    fill.fgColor.rgb = rgb or ""
    cell.fill = fill
    return cell


# ── _cell_has_colors ──────────────────────────────────────────────────────────


class TestCellHasColors:

    def test_rgb_match_returns_true(self):
        cell = make_cell(1, value="Header", rgb="FFFFFF00")
        assert DynamicBomStrategy._cell_has_colors(cell, {"FFFFFF00"}) is True

    def test_rgb_no_match_returns_false(self):
        cell = make_cell(1, value="Header", rgb="FF0000FF")
        assert DynamicBomStrategy._cell_has_colors(cell, {"FFFFFF00"}) is False

    def test_short_color_substring_match(self):
        # Short color patterns (len <= 4) should match as substrings
        cell = make_cell(1, value="Header", rgb="FFFFFF00")
        assert DynamicBomStrategy._cell_has_colors(cell, {"FF00"}) is True

    def test_no_fill_returns_false(self):
        cell = make_cell(1, value="Header", rgb=None)
        assert DynamicBomStrategy._cell_has_colors(cell, {"FFFFFF00"}) is False

    def test_exception_in_fill_returns_false(self):
        """If cell.fill raises, should return False gracefully."""
        from unittest.mock import MagicMock, PropertyMock
        cell = MagicMock()
        cell.column = 1
        cell.value = "X"
        type(cell).fill = PropertyMock(side_effect=Exception("fill error"))
        assert DynamicBomStrategy._cell_has_colors(cell, {"FFFFFF00"}) is False


# ── color_fill mode ───────────────────────────────────────────────────────────


class TestColorFillMode:

    def _make_strategy(self, colors=None, target_column="part_number"):
        cfg = {
            "mode": "color_fill",
            "config": {
                "target_column": target_column,
                "colors": colors or ["FFFFFF00"],
            },
        }
        return DynamicBomStrategy(_make_config(group_detection=cfg))

    def test_target_column_has_matching_color_returns_true(self):
        strategy = self._make_strategy(colors=["FFFFFF00"])
        col_map = {"part_number": 1}
        row = [make_cell(1, value="PN-001", rgb="FFFFFF00")]
        assert strategy.is_group_header(row, col_map) is True

    def test_target_column_different_color_returns_false(self):
        strategy = self._make_strategy(colors=["FFFFFF00"])
        col_map = {"part_number": 1}
        row = [make_cell(1, value="PN-001", rgb="FFFF0000")]
        assert strategy.is_group_header(row, col_map) is False

    def test_target_column_no_fill_returns_false(self):
        strategy = self._make_strategy(colors=["FFFFFF00"])
        col_map = {"part_number": 1}
        row = [make_cell(1, value="PN-001", rgb=None)]
        assert strategy.is_group_header(row, col_map) is False

    def test_target_column_not_in_col_map_returns_false(self):
        strategy = self._make_strategy(colors=["FFFFFF00"], target_column="missing_col")
        col_map = {"part_number": 1}
        row = [make_cell(1, value="PN-001", rgb="FFFFFF00")]
        assert strategy.is_group_header(row, col_map) is False


# ── color_fill_any mode ───────────────────────────────────────────────────────


class TestColorFillAnyMode:

    def _make_strategy(self, colors=None):
        cfg = {
            "mode": "color_fill_any",
            "config": {"colors": colors or ["FFFFFF00"]},
        }
        return DynamicBomStrategy(_make_config(group_detection=cfg))

    def test_pn_column_match_returns_true(self):
        strategy = self._make_strategy(colors=["FFFFFF00"])
        col_map = {"part_number": 1}
        row = [make_cell(1, value="PN-001", rgb="FFFFFF00"), make_cell(2, value="Desc", rgb=None)]
        assert strategy.is_group_header(row, col_map) is True

    def test_any_cell_with_matching_color_returns_true(self):
        strategy = self._make_strategy(colors=["FFFFFF00"])
        col_map = {"part_number": 2}  # pn is col 2, col 1 has the color
        row = [make_cell(1, value="Header", rgb="FFFFFF00"), make_cell(2, value=None, rgb=None)]
        assert strategy.is_group_header(row, col_map) is True

    def test_no_matching_color_anywhere_returns_false(self):
        strategy = self._make_strategy(colors=["FFFFFF00"])
        col_map = {"part_number": 1}
        row = [make_cell(1, value="PN-001", rgb=None), make_cell(2, value="Desc", rgb=None)]
        assert strategy.is_group_header(row, col_map) is False

    def test_row_with_no_values_returns_false(self):
        strategy = self._make_strategy(colors=["FFFFFF00"])
        col_map = {"part_number": 1}
        row = [make_cell(1, value=None, rgb=None)]
        assert strategy.is_group_header(row, col_map) is False


# ── line_number_depth mode ────────────────────────────────────────────────────


class TestLineDepthMode:

    def _make_strategy(self, group_pattern=r"^\d+\.\d+$", line_column="line_number"):
        cfg = {
            "mode": "line_number_depth",
            "config": {
                "line_column": line_column,
                "group_pattern": group_pattern,
            },
        }
        return DynamicBomStrategy(_make_config(group_detection=cfg))

    def test_line_matches_pattern_returns_true(self):
        strategy = self._make_strategy()
        col_map = {"line_number": 1}
        row = [make_cell(1, value="2.3")]
        assert strategy.is_group_header(row, col_map) is True

    def test_line_does_not_match_pattern_returns_false(self):
        strategy = self._make_strategy()
        col_map = {"line_number": 1}
        row = [make_cell(1, value="2.3.4")]  # 3-segment — doesn't match 2-segment pattern
        assert strategy.is_group_header(row, col_map) is False

    def test_line_column_not_in_col_map_returns_false(self):
        strategy = self._make_strategy()
        col_map = {"part_number": 1}  # no line_number key
        row = [make_cell(1, value="1.2")]
        assert strategy.is_group_header(row, col_map) is False

    def test_empty_cell_value_returns_false(self):
        strategy = self._make_strategy()
        col_map = {"line_number": 1}
        row = [make_cell(1, value=None)]
        assert strategy.is_group_header(row, col_map) is False

    def test_custom_pattern_three_segment(self):
        strategy = self._make_strategy(group_pattern=r"^\d+\.\d+\.\d+$")
        col_map = {"line_number": 1}
        row = [make_cell(1, value="1.2.3")]
        assert strategy.is_group_header(row, col_map) is True


# ── value_change mode ─────────────────────────────────────────────────────────


class TestValueChangeMode:

    def _make_strategy(self, watch_column="part_number", condition_column="ext_qty", condition_value=1):
        cfg = {
            "mode": "value_change",
            "config": {
                "watch_column": watch_column,
                "condition_column": condition_column,
                "condition_value": condition_value,
            },
        }
        return DynamicBomStrategy(_make_config(group_detection=cfg))

    def test_value_change_with_correct_condition_returns_true(self):
        strategy = self._make_strategy()
        col_map = {"part_number": 1, "ext_qty": 2}
        row = [make_cell(1, value="PN-002"), make_cell(2, value=1)]
        assert strategy.is_group_header(row, col_map, prev_part="PN-001") is True

    def test_same_part_number_returns_false(self):
        strategy = self._make_strategy()
        col_map = {"part_number": 1, "ext_qty": 2}
        row = [make_cell(1, value="PN-001"), make_cell(2, value=1)]
        assert strategy.is_group_header(row, col_map, prev_part="PN-001") is False

    def test_different_pn_wrong_condition_value_returns_false(self):
        strategy = self._make_strategy(condition_value=1)
        col_map = {"part_number": 1, "ext_qty": 2}
        row = [make_cell(1, value="PN-002"), make_cell(2, value=5)]  # qty=5, not 1
        assert strategy.is_group_header(row, col_map, prev_part="PN-001") is False

    def test_none_prev_part_treats_as_change(self):
        """When prev_part is None, any part number is a change."""
        strategy = self._make_strategy()
        col_map = {"part_number": 1, "ext_qty": 2}
        row = [make_cell(1, value="PN-001"), make_cell(2, value=1)]
        assert strategy.is_group_header(row, col_map, prev_part=None) is True

    def test_missing_watch_column_returns_false(self):
        strategy = self._make_strategy()
        col_map = {"ext_qty": 2}  # no part_number in map
        row = [make_cell(1, value="PN-001"), make_cell(2, value=1)]
        assert strategy.is_group_header(row, col_map, prev_part=None) is False

    def test_invalid_condition_value_treated_as_zero(self):
        strategy = self._make_strategy(condition_value=0)
        col_map = {"part_number": 1, "ext_qty": 2}
        row = [make_cell(1, value="PN-002"), make_cell(2, value="not_a_number")]
        # value="not_a_number" → qty=0, condition_value=0 → True
        assert strategy.is_group_header(row, col_map, prev_part="PN-001") is True
