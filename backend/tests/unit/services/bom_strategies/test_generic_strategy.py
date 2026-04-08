"""
Tests for GenericBomStrategy – fallback parser logic.
"""
import pytest
import openpyxl
from app.services.bom_strategies.generic_strategy import GenericBomStrategy


@pytest.fixture
def strategy():
    return GenericBomStrategy()


class TestGenericStrategyProperties:

    def test_format_id(self, strategy):
        assert strategy.format_id == "generic_first_col"

    def test_column_map_covers_all_vendor_headers(self, strategy):
        col_map = strategy.get_column_map()
        # NetApp headers
        assert col_map["Part Number"] == "part_number"
        assert col_map["Product"] == "product"
        assert col_map["Ext Qty"] == "ext_qty"
        # Dell headers
        assert col_map["Sku"] == "part_number"
        assert col_map["SKU"] == "part_number"
        assert col_map["Quantity"] == "ext_qty"
        assert col_map["Total Selling Price"] == "ext_net_price"
        # HPE headers
        assert col_map["UCID"] == "part_number"
        assert col_map["Total Net Price"] == "ext_net_price"
        # Common
        assert col_map["Description"] == "product"
        assert col_map["Discount"] == "net_discount"
        assert col_map["Category"] == "mod_group"


class TestGenericFindHeaderRow:

    def test_finds_part_number_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=3, column=2, value="Part Number")
        assert strategy.find_header_row(ws) == 3

    def test_finds_sku_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="sku")
        assert strategy.find_header_row(ws) == 1

    def test_finds_ucid_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="UCID")
        assert strategy.find_header_row(ws) == 2

    def test_finds_description_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Description")
        assert strategy.find_header_row(ws) == 1

    def test_finds_qty_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="qty")
        assert strategy.find_header_row(ws) == 1

    def test_returns_none_for_empty_sheet(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert strategy.find_header_row(ws) is None

    def test_returns_none_for_unrecognized_headers(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Foo")
        ws.cell(row=2, column=1, value="Bar")
        assert strategy.find_header_row(ws) is None


class TestGenericIsGroupHeader:

    def _make_row(self, ws, part_number, qty, row_num=1):
        ws.cell(row=row_num, column=1, value=part_number)
        ws.cell(row=row_num, column=2, value=qty)
        return list(ws.iter_rows(min_row=row_num, max_row=row_num))[0]

    def test_new_part_with_qty_1_is_group_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"part_number": 1, "ext_qty": 2}
        row = self._make_row(ws, "SYS-001", 1)
        assert strategy.is_group_header(row, col_map, prev_part=None) is True

    def test_same_part_as_prev_is_not_group_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"part_number": 1, "ext_qty": 2}
        row = self._make_row(ws, "SYS-001", 1)
        assert strategy.is_group_header(row, col_map, prev_part="SYS-001") is False

    def test_qty_greater_than_1_is_not_group_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"part_number": 1, "ext_qty": 2}
        row = self._make_row(ws, "CHILD-001", 10)
        assert strategy.is_group_header(row, col_map, prev_part=None) is False

    def test_empty_part_number_with_qty_1_is_group_header(self, strategy):
        """Empty string is != prev_part(None) and qty==1, so it qualifies."""
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"part_number": 1, "ext_qty": 2}
        row = self._make_row(ws, "", 1)
        assert strategy.is_group_header(row, col_map) is True

    def test_none_part_number_with_qty_1_is_group_header(self, strategy):
        """None coerces to empty string, which differs from prev_part(None)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"part_number": 1, "ext_qty": 2}
        row = self._make_row(ws, None, 1)
        assert strategy.is_group_header(row, col_map) is True

    def test_non_numeric_qty_is_not_group_header(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"part_number": 1, "ext_qty": 2}
        row = self._make_row(ws, "SYS-001", "N/A")
        assert strategy.is_group_header(row, col_map) is False

    def test_missing_qty_column_in_col_map(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"part_number": 1}
        row = self._make_row(ws, "SYS-001", 1)
        # Without ext_qty mapped, qty defaults to 0, so not group header
        assert strategy.is_group_header(row, col_map) is False
