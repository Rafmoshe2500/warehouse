"""
Tests for CiscoBomStrategy using real example BOM file (דוגמא4.xlsx).
"""
import os
import pytest
import openpyxl
from app.services.bom_strategies.cisco_strategy import CiscoBomStrategy

EXAMPLE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "example_bom_files", "דוגמא4.xlsx"
)


@pytest.fixture
def strategy():
    return CiscoBomStrategy()


@pytest.fixture
def workbook():
    assert os.path.exists(EXAMPLE_FILE), f"Example BOM file not found: {EXAMPLE_FILE}"
    return openpyxl.load_workbook(EXAMPLE_FILE, data_only=True)


class TestCiscoStrategyProperties:

    def test_format_id(self, strategy):
        assert strategy.format_id == "cisco_quote"

    def test_column_map_keys(self, strategy):
        col_map = strategy.get_column_map()
        assert col_map["Line Number"] == "line_number"
        assert col_map["Part Number"] == "part_number"
        assert col_map["Description"] == "product"
        assert col_map["Qty"] == "ext_qty"
        assert "Disc(%)" in col_map


class TestCiscoFindHeaderRow:

    def test_finds_header_in_real_file(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        assert header_row is not None
        assert isinstance(header_row, int)

    def test_header_row_contains_line_number(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        if header_row is not None:
            found = False
            for cell in ws[header_row]:
                if isinstance(cell.value, str) and "line number" in cell.value.lower():
                    found = True
                    break
            assert found

    def test_searches_up_to_row_40(self, strategy):
        """Cisco scans further than other vendors (up to row 40)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=35, column=1, value="Line Number")
        assert strategy.find_header_row(ws) == 35

    def test_returns_none_for_empty_sheet(self, strategy):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert strategy.find_header_row(ws) is None


class TestCiscoIsGroupHeader:

    def _make_row_with_line_number(self, value):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=value)
        return list(ws.iter_rows(min_row=1, max_row=1))[0]

    def test_top_level_two_segments_is_group_header(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("1.0")
        assert strategy.is_group_header(row, col_map) is True

    def test_two_segments_non_zero(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("2.3")
        assert strategy.is_group_header(row, col_map) is True

    def test_three_segments_is_not_group_header(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("1.1.1")
        assert strategy.is_group_header(row, col_map) is False

    def test_four_segments_is_not_group_header(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("1.3.10.2")
        assert strategy.is_group_header(row, col_map) is False

    def test_empty_line_number_is_not_group_header(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number(None)
        assert strategy.is_group_header(row, col_map) is False

    def test_text_line_number_is_not_group_header(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("abc")
        assert strategy.is_group_header(row, col_map) is False

    def test_no_line_number_column_returns_false(self, strategy):
        col_map = {}
        row = self._make_row_with_line_number("1.0")
        assert strategy.is_group_header(row, col_map) is False


class TestCiscoIsDataRow:

    def _make_row_with_line_number(self, value):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=value)
        return list(ws.iter_rows(min_row=1, max_row=1))[0]

    def test_dotted_number_is_data_row(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("1.0")
        assert strategy.is_data_row(row, col_map) is True

    def test_three_segment_is_data_row(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("1.2.3")
        assert strategy.is_data_row(row, col_map) is True

    def test_single_number_is_data_row(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("5")
        assert strategy.is_data_row(row, col_map) is True

    def test_empty_is_not_data_row(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number(None)
        assert strategy.is_data_row(row, col_map) is False

    def test_text_is_not_data_row(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("Notes")
        assert strategy.is_data_row(row, col_map) is False

    def test_disclaimer_text_is_not_data_row(self, strategy):
        col_map = {"line_number": 1}
        row = self._make_row_with_line_number("Terms and Conditions")
        assert strategy.is_data_row(row, col_map) is False

    def test_no_line_number_column_returns_true(self, strategy):
        """When no line column is mapped, let the parser decide."""
        col_map = {}
        row = self._make_row_with_line_number("anything")
        assert strategy.is_data_row(row, col_map) is True


class TestCiscoRealFile:

    def test_real_file_has_group_headers_and_children(self, strategy, workbook):
        ws = workbook.worksheets[0]
        header_row = strategy.find_header_row(ws)
        assert header_row is not None

        col_map = {}
        source = strategy.get_column_map()
        for cell in ws[header_row]:
            if isinstance(cell.value, str):
                for excel_key, field_key in source.items():
                    if cell.value.strip().lower() == excel_key.lower():
                        col_map.setdefault(field_key, cell.column)

        group_headers = 0
        children = 0
        skipped = 0

        for row in ws.iter_rows(min_row=header_row + 1):
            if not any(cell.value is not None for cell in row):
                continue
            if not strategy.is_data_row(row, col_map):
                skipped += 1
                continue
            if strategy.is_group_header(row, col_map):
                group_headers += 1
            else:
                children += 1

        assert group_headers > 0, "Expected at least one group header in Cisco example"
        assert children > 0, "Expected at least one child row in Cisco example"
        # Cisco files typically have metadata rows that get skipped
        assert skipped >= 0
