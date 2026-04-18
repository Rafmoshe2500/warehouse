"""
Unit tests for BomTemplateService — preview_excel, validate_template, and CRUD methods.
"""
import io
import json
import pytest
from unittest.mock import AsyncMock, MagicMock, patch
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.services.bom_template_service import BomTemplateService, _slugify


# ── Helpers ────────────────────────────────────────────────────────────────


def make_xlsx(headers: list, rows: list, fill_row: int = None, fill_rgb: str = None) -> bytes:
    """Create an in-memory xlsx with given headers and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for i, row in enumerate(rows, start=2):  # row 2 onwards is data
        ws.append(row)
        if fill_row == i and fill_rgb:
            for cell in ws[i]:
                cell.fill = PatternFill(start_color=fill_rgb, end_color=fill_rgb, fill_type="solid")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _base_template_data(**overrides) -> dict:
    base = {
        "format_id": "test_vendor",
        "vendor_name": "TestVendor",
        "header_detection": {"keyword": "Part Number", "max_scan_rows": 25},
        "column_map": {"Part Number": "part_number", "Description": "product", "Qty": "ext_qty"},
        "group_detection": {"mode": "all_rows", "config": {}},
    }
    base.update(overrides)
    return base


# ── Slugify utility ────────────────────────────────────────────────────────


class TestSlugify:
    def test_simple_name(self):
        assert _slugify("Dell") == "dell"

    def test_spaces_become_underscores(self):
        assert _slugify("Test Vendor") == "test_vendor"

    def test_special_chars_removed(self):
        assert _slugify("HPE/NetApp-2025!") == "hpe_netapp_2025"

    def test_leading_trailing_underscores_stripped(self):
        assert _slugify("  __vendor__  ") == "vendor"


# ── preview_excel ──────────────────────────────────────────────────────────


class TestPreviewExcel:

    def test_returns_sheet_name_and_rows(self):
        file_bytes = make_xlsx(
            headers=["Part Number", "Description", "Qty"],
            rows=[["PN-001", "Widget A", 5], ["PN-002", "Widget B", 3]],
        )
        result = BomTemplateService.preview_excel(file_bytes)

        assert "sheet_name" in result
        assert "rows" in result
        assert "total_sheets" in result
        assert result["total_sheets"] == 1
        assert len(result["rows"]) >= 1  # header row returned

    def test_row_structure_contains_value_and_col(self):
        file_bytes = make_xlsx(
            headers=["Part Number", "Qty"],
            rows=[["PN-001", 10]],
        )
        result = BomTemplateService.preview_excel(file_bytes)

        first_row = result["rows"][0]
        assert "row" in first_row
        assert "cells" in first_row
        first_cell = first_row["cells"][0]
        assert "value" in first_cell
        assert "col" in first_cell

    def test_respects_max_rows(self):
        rows = [[f"PN-{i:03d}", f"Item {i}", i] for i in range(1, 60)]
        file_bytes = make_xlsx(headers=["Part Number", "Description", "Qty"], rows=rows)

        result = BomTemplateService.preview_excel(file_bytes, max_rows=5)
        assert len(result["rows"]) == 5

    def test_includes_fill_color_when_present(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Part Number", "Description"])
        ws.append(["PN-001", "Coloured row"])
        yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        ws["A2"].fill = yellow

        buf = io.BytesIO()
        wb.save(buf)
        file_bytes = buf.getvalue()

        result = BomTemplateService.preview_excel(file_bytes)
        # Find row 2
        data_row = next(r for r in result["rows"] if r["row"] == 2)
        a2_cell = next(c for c in data_row["cells"] if c["col"] == 1)
        assert "fill" in a2_cell
        assert "FFFF00" in a2_cell["fill"] or "FF00" in a2_cell["fill"]

    def test_no_fill_key_when_no_fill(self):
        file_bytes = make_xlsx(
            headers=["Part Number"],
            rows=[["PN-001"]],
        )
        result = BomTemplateService.preview_excel(file_bytes)
        data_row = next(r for r in result["rows"] if r["row"] == 2)
        a2_cell = data_row["cells"][0]
        # No fill or fill key absent/None is fine
        fill_val = a2_cell.get("fill")
        # The service returns '00000000' for no-fill cells (transparent/white)
        assert fill_val is None or fill_val in ("00000000", "FFFFFFFF", "") or "fill" not in a2_cell

    def test_invalid_file_raises(self):
        with pytest.raises(Exception):
            BomTemplateService.preview_excel(b"not an xlsx file at all!!")

    def test_datetime_value_converted_to_isoformat(self):
        from datetime import datetime, timezone
        wb = Workbook()
        ws = wb.active
        ws.append(["Date"])
        ws["A2"] = datetime(2024, 6, 15, 12, 0, 0)
        ws["A2"].number_format = "YYYY-MM-DD"
        buf = io.BytesIO()
        wb.save(buf)
        result = BomTemplateService.preview_excel(buf.getvalue())
        data_row = next(r for r in result["rows"] if r["row"] == 2)
        val = data_row["cells"][0]["value"]
        assert isinstance(val, str)  # converted to isoformat string


# ── validate_template ──────────────────────────────────────────────────────


class TestValidateTemplate:

    def test_valid_config_returns_valid_true(self):
        file_bytes = make_xlsx(
            headers=["Part Number", "Description", "Qty"],
            rows=[["PN-001", "Widget", 5], ["PN-002", "Gadget", 3]],
        )
        template = _base_template_data()
        result = BomTemplateService.validate_template(template, file_bytes)

        assert result["valid"] is True
        assert result["error"] is None
        assert result["groups_count"] >= 1
        assert "mapped_columns" in result
        assert "part_number" in result["mapped_columns"]

    def test_keyword_not_found_returns_valid_false(self):
        file_bytes = make_xlsx(
            headers=["SKU", "Name", "Count"],
            rows=[["SKU-001", "Thing", 1]],
        )
        template = _base_template_data(
            header_detection={"keyword": "part number", "max_scan_rows": 5}
        )
        result = BomTemplateService.validate_template(template, file_bytes)

        assert result["valid"] is False
        assert "keyword" in result["error"].lower() or "header" in result["error"].lower()
        assert result["groups_count"] == 0

    def test_part_number_column_not_mappable_returns_valid_false(self):
        file_bytes = make_xlsx(
            headers=["Part Number", "Description", "Qty"],
            rows=[["PN-001", "Widget", 5]],
        )
        # column_map has no entry mapping to "part_number"
        template = _base_template_data(
            column_map={"Description": "product", "Qty": "ext_qty"}
        )
        result = BomTemplateService.validate_template(template, file_bytes)

        assert result["valid"] is False
        assert "part_number" in result["error"].lower()

    def test_returns_groups_count_and_sample(self):
        file_bytes = make_xlsx(
            headers=["Part Number", "Description", "Qty"],
            rows=[
                ["GRP-001", "Group One", 1],
                ["CHILD-001", "Child One", 3],
                ["GRP-002", "Group Two", 1],
                ["CHILD-002", "Child Two", 2],
            ],
        )
        template = _base_template_data()
        result = BomTemplateService.validate_template(template, file_bytes)

        assert result["valid"] is True
        assert result["groups_count"] >= 1
        assert result["sample_group"] is not None

    def test_validate_with_data_row_filter(self):
        """Only rows matching the filter pattern should be parsed."""
        file_bytes = make_xlsx(
            headers=["Part Number", "Line", "Description"],
            rows=[
                ["PN-001", "1.1", "Group item"],
                ["PN-002", "skip_me", "Non-data"],
                ["PN-003", "2.1", "Another group"],
            ],
        )
        template = _base_template_data(
            column_map={"Part Number": "part_number", "Line": "line_number", "Description": "product"},
            data_row_filter={"column": "line_number", "pattern": r"^\d+\.\d+$"},
        )
        result = BomTemplateService.validate_template(template, file_bytes)
        assert result["valid"] is True
        # Only rows with "line_number" matching \d+.\d+ should be counted
        assert result["groups_count"] == 2

    def test_returns_header_row_index(self):
        file_bytes = make_xlsx(
            headers=["Part Number", "Description"],
            rows=[["PN-001", "Widget"]],
        )
        template = _base_template_data()
        result = BomTemplateService.validate_template(template, file_bytes)

        assert result["valid"] is True
        assert result["header_row"] == 1  # header is the first row

    def test_header_not_in_first_row(self):
        """Header keyword found in row 3."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Company Info", ""])
        ws.append(["Date", "2024"])
        ws.append(["Part Number", "Description", "Qty"])
        ws.append(["PN-001", "Item", 1])
        buf = io.BytesIO()
        wb.save(buf)

        template = _base_template_data(
            header_detection={"keyword": "Part Number", "max_scan_rows": 10}
        )
        result = BomTemplateService.validate_template(template, buf.getvalue())

        assert result["valid"] is True
        assert result["header_row"] == 3


# ── CRUD methods (async, with mocked repository) ──────────────────────────


class TestBomTemplateServiceCRUD:

    @pytest.fixture
    def mock_repo(self):
        repo = AsyncMock()
        repo._format_doc = lambda d: {**d, "id": str(d.get("_id", "test_id"))}
        return repo

    @pytest.fixture
    def service(self, mock_repo):
        svc = BomTemplateService.__new__(BomTemplateService)
        svc.repo = mock_repo
        return svc

    @pytest.mark.asyncio
    async def test_create_template_success(self, service, mock_repo):
        mock_repo.get_by_format_id.return_value = None
        created_doc = {
            "_id": "abc123",
            "vendor_name": "NewVendor",
            "format_id": "newvendor",
            "is_active": True,
        }
        mock_repo.create.return_value = created_doc

        result = await service.create_template(
            data={
                "vendor_name": "NewVendor",
                "column_map": {"Part Number": "part_number"},
                "header_detection": {"keyword": "Part Number", "max_scan_rows": 10},
                "group_detection": {"mode": "all_rows", "config": {}},
            },
            username="admin",
        )
        mock_repo.create.assert_called_once()
        assert result is not None

    @pytest.mark.asyncio
    async def test_create_template_duplicate_format_id_raises(self, service, mock_repo):
        mock_repo.get_by_format_id.return_value = {"format_id": "newvendor"}

        with pytest.raises(ValueError, match="already exists"):
            await service.create_template(
                data={
                    "vendor_name": "NewVendor",
                    "column_map": {"Part Number": "part_number"},
                    "header_detection": {"keyword": "Part Number", "max_scan_rows": 10},
                    "group_detection": {"mode": "all_rows", "config": {}},
                },
                username="admin",
            )

    @pytest.mark.asyncio
    async def test_create_template_empty_vendor_name_raises(self, service, mock_repo):
        with pytest.raises((ValueError, KeyError)):
            await service.create_template(
                data={
                    "vendor_name": "!!!",  # slugifies to empty string
                    "column_map": {"Part Number": "part_number"},
                    "header_detection": {"keyword": "Part Number", "max_scan_rows": 10},
                    "group_detection": {"mode": "all_rows", "config": {}},
                },
                username="admin",
            )

    @pytest.mark.asyncio
    async def test_update_template_not_found_returns_none(self, service, mock_repo):
        mock_repo.get_by_id.return_value = None

        result = await service.update_template("nonexistent_id", {"vendor_name": "X"}, "admin")
        assert result is None

    @pytest.mark.asyncio
    async def test_update_template_success(self, service, mock_repo):
        existing = {"_id": "id1", "vendor_name": "OldName", "is_active": True}
        updated = {"_id": "id1", "vendor_name": "NewName", "is_active": True}
        mock_repo.get_by_id.return_value = existing
        mock_repo.update.return_value = updated

        result = await service.update_template("id1", {"vendor_name": "NewName"}, "admin")
        mock_repo.update.assert_called_once()
        assert result is not None

    @pytest.mark.asyncio
    async def test_update_template_no_fields_returns_existing(self, service, mock_repo):
        existing = {"_id": "id1", "vendor_name": "Same"}
        mock_repo.get_by_id.return_value = existing

        result = await service.update_template("id1", {}, "admin")
        # update should NOT be called since no fields to update
        mock_repo.update.assert_not_called()

    @pytest.mark.asyncio
    async def test_delete_template_not_found_returns_none(self, service, mock_repo):
        mock_repo.get_by_id.return_value = None

        result = await service.delete_template("nonexistent")
        assert result is None

    @pytest.mark.asyncio
    async def test_delete_template_deactivates(self, service, mock_repo):
        existing = {"_id": "id1", "vendor_name": "ToDelete", "is_builtin": False}
        deactivated = {"_id": "id1", "vendor_name": "ToDelete", "is_active": False}
        mock_repo.get_by_id.return_value = existing
        mock_repo.deactivate.return_value = deactivated

        result = await service.delete_template("id1")
        mock_repo.deactivate.assert_called_once_with("id1")

    @pytest.mark.asyncio
    async def test_delete_builtin_template_also_deactivates(self, service, mock_repo):
        existing = {"_id": "id1", "vendor_name": "BuiltIn", "is_builtin": True}
        deactivated = {"_id": "id1", "vendor_name": "BuiltIn", "is_active": False}
        mock_repo.get_by_id.return_value = existing
        mock_repo.deactivate.return_value = deactivated

        await service.delete_template("id1")
        mock_repo.deactivate.assert_called_once_with("id1")
