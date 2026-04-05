import pytest
import io
import openpyxl
from unittest.mock import patch, MagicMock
from app.services.bom_service import BomService, FORMAT_NETAPP, FORMAT_DELL, FORMAT_HPE

@pytest.fixture
def bom_service(mock_mongodb):
    return BomService()

def create_mock_excel(fmt: str) -> bytes:
    """Create a minimal mock BOM Excel file based on format in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if fmt == FORMAT_NETAPP:
        # Headers at row 5
        headers = ["Part Number", "Product", "Ext Qty", "Ext Net Price"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=5, column=col, value=h)
            
        # Group Header (Yellow Part Number)
        cell = ws.cell(row=6, column=1, value="SYS-1")
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        ws.cell(row=6, column=2, value="Main System Product")
        ws.cell(row=6, column=3, value=1)
        ws.cell(row=6, column=4, value=5000)
        
        # Child Item
        ws.cell(row=7, column=1, value="DRV-1")
        ws.cell(row=7, column=2, value="Drive 1TB")
        ws.cell(row=7, column=3, value=10)
        ws.cell(row=7, column=4, value=1000)
        
    elif fmt == FORMAT_DELL:
        headers = ["Sku", "Description", "Qty", "Total Selling Price"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
            
        # Group Header (Dell Yellow SKU)
        cell = ws.cell(row=2, column=1, value="DELL-SYS")
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFFFF00", fill_type="solid")
        ws.cell(row=2, column=2, value="Dell Server")
        ws.cell(row=2, column=3, value=2)
        ws.cell(row=2, column=4, value=4000)
        
        # Child Item
        ws.cell(row=3, column=1, value="DELL-RAM")
        ws.cell(row=3, column=2, value="16GB RAM")
        ws.cell(row=3, column=3, value=4)
        ws.cell(row=3, column=4, value=400)
        
    elif fmt == FORMAT_HPE:
        headers = ["UCID", "Description", "Qty", "Total Net Price"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=h)
            
        # HPE items are generally all main items (no children hierarchy out of the box)
        ws.cell(row=3, column=1, value="HPE-1")
        ws.cell(row=3, column=2, value="HPE Block")
        ws.cell(row=3, column=3, value=1)
        ws.cell(row=3, column=4, value=2000)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

class TestBomService:

    def test_parse_excel_netapp(self, bom_service):
        """Test parse_excel against simulated NetApp structure."""
        file_bytes = create_mock_excel(FORMAT_NETAPP)
        groups, all_part_numbers, extracts = bom_service.parse_excel(file_bytes, fmt=FORMAT_NETAPP)
        
        assert len(groups) == 1
        main = groups[0]["main"]
        assert main["part_number"] == "SYS-1"
        assert main["ext_net_price"] == 5000.0
        
        # Children
        assert len(groups[0]["children"]) == 1
        child = groups[0]["children"][0]
        assert child["part_number"] == "DRV-1"
        assert child["ext_net_price"] == 1000.0
        
        # Total net includes children net pricing (5000 is ignored if there are children in logic context)
        # Wait, the logic for total net price adds up children if they exist:
        assert groups[0]["total_net_price"] == 1000.0
        
        assert "SYS-1" in all_part_numbers
        assert "DRV-1" in all_part_numbers

    def test_parse_excel_dell(self, bom_service):
        """Test parse_excel against simulated Dell structure."""
        file_bytes = create_mock_excel(FORMAT_DELL)
        groups, all_pns, extracts = bom_service.parse_excel(file_bytes, fmt=FORMAT_DELL)
        
        assert len(groups) == 1
        assert groups[0]["main"]["part_number"] == "DELL-SYS"
        assert groups[0]["children"][0]["part_number"] == "DELL-RAM"

    def test_parse_excel_hpe(self, bom_service):
        """Test parse_excel against simulated HPE structure."""
        file_bytes = create_mock_excel(FORMAT_HPE)
        groups, all_pns, extracts = bom_service.parse_excel(file_bytes, fmt=FORMAT_HPE)
        
        # HPE makes everything a group header (main item) without children
        assert len(groups) == 1
        assert groups[0]["main"]["part_number"] == "HPE-1"
        assert len(groups[0]["children"]) == 0
        # If no children, total_net_price matches main
        assert groups[0]["total_net_price"] == 2000.0

    @pytest.mark.asyncio
    async def test_enrich_and_scan_bom(self, bom_service):
        """Mock out classification call and test entire sequence."""
        file_bytes = create_mock_excel(FORMAT_NETAPP)
        
        # We need to mock 'classify_batch' because it calls AI during scan_bom's ML fallback
        with patch("app.ai.classifier.classify_batch") as mock_classify:
            mock_classify.return_value = [
                {"label": "Fake AI Label", "category": "other", "description_he": "Fake", "confidence": 0.99, "low_confidence": False},
                {"label": "Fake AI Label 2", "category": "server", "description_he": "Fake2", "confidence": 0.8, "low_confidence": False}
            ]
            
            # Use bom_service directly. Note: scan_bom actually intercepts classification internally twice
            # once inside enrich_groups and once later in scan_bom loop.
            res = await bom_service.scan_bom(file_bytes, fmt=FORMAT_NETAPP)
            
            assert res["total_groups"] == 1
            assert len(res["unknown_parts"]) == 2
            assert res["unknown_parts"][0]["ai_label"] == "Fake AI Label"
            
            # Validate catalog dictionary exists on main and child items
            assert "catalog" in res["groups"][0]["main"]
            assert "catalog" in res["groups"][0]["children"][0]
